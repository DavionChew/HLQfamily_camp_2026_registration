#!/usr/bin/env python3
"""Create the attendee list template (Excel) that the church fills in.
Columns match the Google Sheet 'Attendees' tab exactly, so the generated
import file can be pasted straight in. Run:  python3 make_template.py
"""
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment

COLS = ['ID','Token','Name','Phone','Emergency','Role','Group','BusTo','BusBack','RoomGroup','Room','RoomNote','Notes']
HELP = {
 'ID':'Leave blank — auto-generated (C001, C002…)',
 'Token':'Leave blank — auto-generated security code',
 'Name':'Full name (Chinese or English) — REQUIRED',
 'Phone':'Contact number',
 'Emergency':'紧急联络人 emergency contact (name + phone)',
 'Role':'Attendee / Organiser / Leader (used for counts; not displayed)',
 'Group':'组别 / 营内分组 — the camp activity group (follows them to 灵修 etc.)',
 'BusTo':'Y if taking the church bus TO the venue, else N',
 'BusBack':'Y if taking the bus BACK to church, else N',
 'RoomGroup':'房间分组 — same code = same room (e.g. R01). Fill BEFORE camp.',
 'Room':'LEAVE BLANK — actual room number auto-filled at check-in (3pm)',
 'RoomNote':'e.g. gender block, special needs',
 'Notes':'allergies, dietary, etc.',
}

wb = Workbook()
ws = wb.active; ws.title = 'Attendees'

hdr_fill = PatternFill('solid', fgColor='1A73E8'); hdr_font = Font(bold=True, color='FFFFFF')
for i, c in enumerate(COLS, 1):
    cell = ws.cell(1, i, c); cell.fill = hdr_fill; cell.font = hdr_font
    ws.cell(2, i, HELP[c]).font = Font(italic=True, size=9, color='888888')
    ws.column_dimensions[cell.column_letter].width = max(12, len(HELP[c]) // 2 + 6)

# example rows
examples = [
 ['', '', '陈大文 David Chen', '0123456789', '母亲 Mum 0111', 'Attendee', 'A 组', 'Y', 'Y', 'R01', '', '', ''],
 ['', '', '林美丽 Mary Lim',   '0129876543', '配偶 Spouse 0122', 'Organiser','B 组','N','N','R02','', '', '负责报到'],
]
for r, row in enumerate(examples, 3):
    for ci, v in enumerate(row, 1):
        ws.cell(r, ci, v)

# dropdowns
def add_dv(col_letter, values):
    dv = DataValidation(type='list', formula1='"%s"' % ','.join(values), allow_blank=True)
    ws.add_data_validation(dv); dv.add('%s3:%s1000' % (col_letter, col_letter))
add_dv('F', ['Attendee', 'Organiser', 'Leader'])   # Role
add_dv('H', ['Y', 'N'])                            # BusTo
add_dv('I', ['Y', 'N'])                            # BusBack

ws.freeze_panes = 'A3'

# instructions sheet
info = wb.create_sheet('READ ME')
lines = [
 'HOW TO USE THIS TEMPLATE',
 '',
 '1. Fill ONE row per attendee on the "Attendees" tab (row 3 onwards).',
 '   You can delete the two example rows.',
 '2. Leave ID and Token BLANK — they are generated automatically.',
 '3. Name is required. Everything else is optional but useful.',
 '4. Role = Organiser for your core team (so they show as 同工).',
 '5. BusTo / BusBack = Y only for people using the church bus.',
 '6. RoomGroup: give everyone sharing a room the SAME code (e.g. R01).',
 '   Leave Room blank — you type each real room number on the Rooms tab',
 '   at 3pm, and it auto-fills to every member when you scan them in.',
 '',
 'WHEN DONE: send this file back. The generator will:',
 '  • assign each person an ID + security token',
 '  • print all the name tags (front = name, back = QR code)',
 '  • produce the import file to paste into the Google Sheet.',
]
for r, t in enumerate(lines, 1):
    c = info.cell(r, 1, t)
    if r == 1: c.font = Font(bold=True, size=14)
info.column_dimensions['A'].width = 70

wb.save('Attendee_List_Template.xlsx')
print('wrote Attendee_List_Template.xlsx')
