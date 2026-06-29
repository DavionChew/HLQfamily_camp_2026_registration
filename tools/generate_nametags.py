#!/usr/bin/env python3
"""
Generate printable name tags (front = name, back = QR) + a Google-Sheet import file.

Usage:
    python3 generate_nametags.py  Attendee_List_Template.xlsx
    python3 generate_nametags.py  my_attendees.csv  --event "GBC Camp 2026"

Outputs (next to the input):
    nametags.pdf          -> print DOUBLE-SIDED, flip on LONG edge, then cut
    attendees_import.xlsx -> paste columns A:K into the Sheet 'Attendees' tab

QR payload = "<ID>-<Token>"  e.g.  C001-7F3K
"""
import sys, io, csv, random, argparse
import qrcode
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.colors import HexColor
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont

pdfmetrics.registerFont(UnicodeCIDFont('STSong-Light'))   # built-in CJK font (Chinese OK)
FONT = 'STSong-Light'

PROFILE = ['ID','Token','Name','Phone','Emergency','Role','Group','CampGroup','BusTo','BusBack','RoomGroup','Room','RoomNote','Notes']
TOKEN_ALPHABET = 'ABCDEFGHJKLMNPQRSTUVWXYZ23456789'   # no 0/O/1/I confusion
# Public schedule/info page the QR opens when scanned by a normal phone camera.
DEFAULT_INFO_URL = 'https://davionchew.github.io/HLQfamily_camp_2026_registration/info.html'
INFO_URL = DEFAULT_INFO_URL
BLUE = HexColor('#1A73E8'); DARK = HexColor('#202124'); GRAY = HexColor('#9aa0a6')

# ---- card / page geometry ----
PAGE_W, PAGE_H = A4
CARD_W, CARD_H = 90*mm, 120*mm
COLS, ROWS = 2, 2
MARGIN_X = (PAGE_W - COLS*CARD_W) / 2
MARGIN_Y = (PAGE_H - ROWS*CARD_H) / 2


def rand_token(n=4):
    return ''.join(random.choice(TOKEN_ALPHABET) for _ in range(n))


def read_rows(path):
    rows = []
    if path.lower().endswith('.csv'):
        with open(path, newline='', encoding='utf-8-sig') as f:
            for d in csv.DictReader(f):
                rows.append({k: (v or '') for k, v in d.items()})
    else:
        wb = load_workbook(path)
        ws = wb['Attendees'] if 'Attendees' in wb.sheetnames else wb.active
        headers = [str(c.value).strip() if c.value is not None else '' for c in ws[1]]
        for rd in ws.iter_rows(min_row=2, values_only=True):
            d = {headers[i]: ('' if i >= len(rd) or rd[i] is None else rd[i]) for i in range(len(headers))}
            rows.append(d)
    # keep only rows with a Name; drop the italic help row / example rows that have no real name
    out = []
    for d in rows:
        name = str(d.get('Name', '')).strip()
        if name and not name.lower().startswith('full name'):
            out.append(d)
    return out


def assign_ids(rows):
    used = set(str(d.get('ID', '')).strip().upper() for d in rows if str(d.get('ID', '')).strip())
    seq = 1
    for d in rows:
        if not str(d.get('ID', '')).strip():
            while ('C%03d' % seq) in used:
                seq += 1
            d['ID'] = 'C%03d' % seq; used.add(d['ID']); seq += 1
        d['ID'] = str(d['ID']).strip().upper()
        if not str(d.get('Token', '')).strip():
            d['Token'] = rand_token()
        d['Token'] = str(d['Token']).strip().upper()
    return rows


def qr_reader(data):
    q = qrcode.QRCode(border=1, box_size=10, error_correction=qrcode.constants.ERROR_CORRECT_M)
    q.add_data(data); q.make(fit=True)
    img = q.make_image(fill_color='black', back_color='white').convert('RGB')
    bio = io.BytesIO(); img.save(bio, format='PNG'); bio.seek(0)
    return ImageReader(bio)


def fit_font(c, text, max_w, start, min_size=9):
    s = start
    while s > min_size and c.stringWidth(text, FONT, s) > max_w:
        s -= 1
    return s


def card_xy(col, row):
    x = MARGIN_X + col*CARD_W
    y = PAGE_H - MARGIN_Y - (row+1)*CARD_H
    return x, y


def crop_marks(c, x, y):
    m = 4*mm; c.setStrokeColor(GRAY); c.setLineWidth(0.3)
    for (cx, cy) in [(x, y), (x+CARD_W, y), (x, y+CARD_H), (x+CARD_W, y+CARD_H)]:
        c.line(cx-m, cy, cx-m+2.5*mm, cy); c.line(cx, cy-m, cx, cy-m+2.5*mm)


def draw_front(c, x, y, att, event):
    c.setStrokeColor(HexColor('#d0d4da')); c.setLineWidth(0.5)
    c.rect(x, y, CARD_W, CARD_H)
    # top band
    band = 22*mm
    c.setFillColor(BLUE); c.rect(x, y+CARD_H-band, CARD_W, band, fill=1, stroke=0)
    c.setFillColor(HexColor('#ffffff'))
    es = fit_font(c, event, CARD_W-14*mm, 15)
    c.setFont(FONT, es); c.drawCentredString(x+CARD_W/2, y+CARD_H-band+7*mm, event)
    # name
    name = str(att['Name'])
    ns = fit_font(c, name, CARD_W-12*mm, 30, 14)
    c.setFillColor(DARK); c.setFont(FONT, ns)
    c.drawCentredString(x+CARD_W/2, y+CARD_H-band-20*mm, name)
    # group
    grp = str(att.get('Group', '') or '')
    if grp:
        c.setFillColor(GRAY); c.setFont(FONT, fit_font(c, grp, CARD_W-16*mm, 14))
        c.drawCentredString(x+CARD_W/2, y+CARD_H-band-32*mm, grp)
    # role badge
    role = str(att.get('Role', '') or '')
    if role.lower().startswith('organ'):
        bw, bh = 34*mm, 9*mm; bx = x+(CARD_W-bw)/2; by = y+22*mm
        c.setFillColor(HexColor('#5b2a86')); c.roundRect(bx, by, bw, bh, 3*mm, fill=1, stroke=0)
        c.setFillColor(HexColor('#ffffff')); c.setFont(FONT, 11)
        c.drawCentredString(x+CARD_W/2, by+2.6*mm, '同工 ORGANISER')
    # id footer
    c.setFillColor(GRAY); c.setFont(FONT, 9)
    c.drawCentredString(x+CARD_W/2, y+8*mm, str(att['ID']))
    crop_marks(c, x, y)


def draw_back(c, x, y, att):
    c.setStrokeColor(HexColor('#d0d4da')); c.setLineWidth(0.5)
    c.rect(x, y, CARD_W, CARD_H)
    idtok = '%s-%s' % (att['ID'], att['Token'])
    # QR encodes the info-page URL (so a normal phone camera shows the schedule),
    # with ?id= that the organiser scanner reads for check-in. No URL -> plain id-token.
    if INFO_URL:
        qr_data = INFO_URL + ('&' if '?' in INFO_URL else '?') + 'id=' + idtok
    else:
        qr_data = idtok
    qsize = 58*mm
    c.drawImage(qr_reader(qr_data), x+(CARD_W-qsize)/2, y+CARD_H-qsize-20*mm, qsize, qsize)
    c.setFillColor(DARK); c.setFont(FONT, 13)
    c.drawCentredString(x+CARD_W/2, y+32*mm, idtok)                       # short code (for manual entry)
    c.setFillColor(GRAY); c.setFont(FONT, fit_font(c, str(att['Name']), CARD_W-12*mm, 12))
    c.drawCentredString(x+CARD_W/2, y+24*mm, str(att['Name']))
    c.setFillColor(BLUE); c.setFont(FONT, 10)
    c.drawCentredString(x+CARD_W/2, y+15*mm, '扫描看节目表 Scan for schedule')
    c.setFillColor(GRAY); c.setFont(FONT, 8)
    c.drawCentredString(x+CARD_W/2, y+9*mm, '同工扫描即报到 · Organisers: scan to check in')
    crop_marks(c, x, y)


def build_pdf(rows, out_pdf, event):
    c = canvas.Canvas(out_pdf, pagesize=A4)
    per = COLS*ROWS
    for i in range(0, len(rows), per):
        chunk = rows[i:i+per]
        # FRONT page
        for j, att in enumerate(chunk):
            col, row = j % COLS, j // COLS
            x, y = card_xy(col, row); draw_front(c, x, y, att, event)
        c.showPage()
        # BACK page — mirror columns for duplex long-edge flip
        for j, att in enumerate(chunk):
            col, row = j % COLS, j // COLS
            x, y = card_xy(COLS-1-col, row); draw_back(c, x, y, att)
        c.showPage()
    c.save()


def write_import(rows, out_xlsx):
    wb = Workbook(); ws = wb.active; ws.title = 'Attendees'
    ws.append(PROFILE)
    for d in rows:
        ws.append([d.get(k, '') for k in PROFILE])
    wb.save(out_xlsx)


def main():
    global INFO_URL
    ap = argparse.ArgumentParser()
    ap.add_argument('input')
    ap.add_argument('--event', default='Church Camp 2026')
    ap.add_argument('--pdf', default='nametags.pdf')
    ap.add_argument('--import-file', default='attendees_import.xlsx')
    ap.add_argument('--info-url', default=DEFAULT_INFO_URL,
                    help='Public info/schedule page the QR opens. Use "" to encode plain ID-Token only.')
    a = ap.parse_args()
    INFO_URL = a.info_url

    rows = assign_ids(read_rows(a.input))
    if not rows:
        print('No attendees with a Name found in', a.input); sys.exit(1)
    build_pdf(rows, a.pdf, a.event)
    write_import(rows, a.import_file)
    print('Attendees: %d' % len(rows))
    print('Wrote %s  (%d pages incl. backs)' % (a.pdf, (len(rows)+3)//4*2))
    print('Wrote %s' % a.import_file)


if __name__ == '__main__':
    main()
